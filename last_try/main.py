import ezdxf
import openpyxl
import re
from collections import defaultdict

def extract_rcc_data(dxf_path, excel_path, proximity_thresh=5000):
    # Load DXF document
    try:
        doc = ezdxf.readfile(dxf_path)
        print(f"Successfully loaded DXF file: {dxf_path}")
    except (IOError, ezdxf.DXFStructureError) as e:
        print(f"Error reading DXF: {e}")
        return False
    
    msp = doc.modelspace()

    # Prepare Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beam Data"
    
    # Column headers
    headers = [
        "BEAM NO", "WIDTH (mm)", "DEPTH (mm)", "LEVEL",
        # … the rest of your headers …
    ]
    ws.append(headers)
    
    # Patterns (level now only 1,2,3)
    patterns = {
        'beam_no': re.compile(r'\bB[0-9]+[A-Za-z]?\b'),
        'section': re.compile(r'(\d{3,4})\s*[xX]\s*(\d{3,4})'),
        'level': re.compile(r'\b(?:L|LEVEL)[:\-]?\s*([123])\b', re.IGNORECASE),
        'rebar': re.compile(r'(\d+)\s*[-\#]\s*(\d+)\s*(?:mm|Ø)?'),
        'stirrup': re.compile(r'(\d+)\s*(?:mm|Ø)?\s*@\s*(\d+)\s*(?:mm|cm|c/c)?'),
        'bent_up': re.compile(r'(?:BENT\s*UP|EXTRA\s*TOP)\s*[:\-]?\s*(\d+)\s*[-\#]\s*(\d+)', re.IGNORECASE)
    }
    
    # Collect all TEXT/MTEXT with positions
    text_entities = []
    for e in msp:
        if e.dxftype() in ("TEXT", "MTEXT"):
            txt = e.dxf.text.replace('\n',' ').strip()
            text_entities.append((txt, (e.dxf.insert.x, e.dxf.insert.y)))
    print(f"Found {len(text_entities)} text entities")
    
    # Build beam_data
    beam_data = defaultdict(dict)
    for text, pos in text_entities:
        # Beam ID
        m = patterns['beam_no'].search(text)
        if m:
            bid = m.group()
            beam_data.setdefault(bid, {})['pos'] = pos
            beam_data[bid].setdefault('source', []).append(text)
        
        # Section dims
        m = patterns['section'].search(text)
        if m:
            w,d = m.groups()
            # assign to nearest beam if not inline
            # … your existing proximity logic …
            # beam_data[bid]['width'], ['depth'] = w, d
        
        # Level
        m = patterns['level'].search(text)
        if m:
            lvl = int(m.group(1))
            # find nearest beam within threshold
            nearest, mind = None, float('inf')
            for b,data in beam_data.items():
                if 'pos' not in data: continue
                dx = pos[0]-data['pos'][0]; dy = pos[1]-data['pos'][1]
                dist = (dx*dx+dy*dy)**0.5
                if dist < mind:
                    mind, nearest = dist, b
            if nearest and mind < proximity_thresh:
                beam_data[nearest]['level'] = lvl
                print(f" → Level {lvl} for {nearest} (d={mind:.1f})")
        
        # … existing rebar and stirrup extraction …
    
    # Write to Excel
    for row_idx, (bid, data) in enumerate(sorted(beam_data.items()), start=2):
        ws.cell(row=row_idx, column=1, value=bid)
        ws.cell(row=row_idx, column=2, value=int(data.get('width', 0)))
        ws.cell(row=row_idx, column=3, value=int(data.get('depth', 0)))
        # Only write level if present
        if 'level' in data:
            ws.cell(row=row_idx, column=4, value=data['level'])
        # … write remaining columns …
    
    wb.save(excel_path)
    print(f"Saved {len(beam_data)} beams to {excel_path}")
    return True

if __name__ == '__main__':
    extract_rcc_data('TRIAL1_BS.dxf', 'RCC_Beam_Data_Enhanced.xlsx')
