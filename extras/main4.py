import ezdxf
import openpyxl
import re
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict

def extract_rcc_data(dxf_path, excel_path, proximity_thresh=5000, debug=False):
    # Read DXF
    try:
        doc = ezdxf.readfile(dxf_path)
    except Exception as e:
        print(f"Error: {e}")
        return
    msp = doc.modelspace()

    # Initialize Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beam Data"
    headers = [
        "BEAM NO","WIDTH","DEPTH","LEVEL",
        "LEFT SPACE STIRRUPS","MID SPACE STIRRUPS","RIGHT SPACE STIRRUPS",
        "SHEAR STIRUP LEG","SHEAR STIRRUPS DIA (L)","SHEAR STIRRUPS DIA (M)","SHEAR STIRRUPS DIA (R)",
        "SHEAR STIRRUP NUMBER","EXTRA STIRRUP NUMBER","EXTRA STIRRUP DIA",
        "HORI LINK DIA","STIRRUPS ID CONTINUOUS END","DISCONTINUOUS END","ATTACH MASTER ID"
    ]
    # Style
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(bold=True)
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.fill = header_fill; cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 20

    # Patterns based on screenshot notation
    beam_no_pat = re.compile(r"\bB\d+[A-Za-z]?\b")
    dim_pat = re.compile(r"(\d+)\s*[xxX]\s*(\d+)")
    lvl_pat = re.compile(r"\+?(\d+\.\d+)" )  # captures +84.05 etc
    # Stirrup categories
    patterns = {
        "LEFT SPACE STIRRUPS": re.compile(r"2-\d+\(C\)"),
        "MID SPACE STIRRUPS": re.compile(r"3-\d+\(T\)"),
        "RIGHT SPACE STIRRUPS": re.compile(r"2-\d+\(T\)"),
        "SHEAR STIRUP LEG": re.compile(r"(\d)-\d+\(T\)"),
        "SHEAR STIRRUPS DIA (L)": re.compile(r"\+1-\d+\(C\)"),
        "SHEAR STIRRUPS DIA (M)": re.compile(r"\+1-\d+\(T\)"),
        "SHEAR STIRRUPS DIA (R)": re.compile(r"\+1-\d+\(C\)\s*ALL"),
        "SHEAR STIRRUP NUMBER": re.compile(r"ALL/st8"),
        "EXTRA STIRRUP NUMBER": re.compile(r"\d+ st8"),
        "EXTRA STIRRUP DIA": re.compile(r"1-\d+\(T\)"),
        "HORI LINK DIA": re.compile(r"230"),
        "STIRRUPS ID CONTINUOUS END": re.compile(r"ALL"),
        "DISCONTINUOUS END": re.compile(r"st6"),
        "ATTACH MASTER ID": re.compile(r"B\d+[A-Za-z]?a?-B\d+[A-Za-z]?a?")
    }

    # Collect text entities
    texts = [(e.dxf.text.strip(), e.dxf.insert.x, e.dxf.insert.y) for e in msp if e.dxftype() in ("TEXT","MTEXT")]
    if debug:
        for t,x,y in texts:
            print(f"Text='{t}' @({x:.1f},{y:.1f})")

    # Data store
    beam_data = defaultdict(lambda: {h: None for h in headers})

    # First, locate beams
    for txt,x,y in texts:
        m = beam_no_pat.search(txt)
        if not m: continue
        b = m.group(0)
        beam_data[b]["BEAM NO"] = b
        beam_data[b]["pos"] = (x,y)
        # inline dims
        d = dim_pat.search(txt)
        if d:
            beam_data[b]["WIDTH"], beam_data[b]["DEPTH"] = d.groups()

    # Helper: nearest
    def nearest(x,y):
        best,bd = None,1e6
        for b,data in beam_data.items():
            if "pos" not in data: continue
            dx,dy = x-data["pos"][0], y-data["pos"][1]
            dist = (dx*dx+dy*dy)**0.5
            if dist<bd:
                best,bd = b,dist
        return best if bd<proximity_thresh else None

    # Second pass: assign other attributes
    for txt,x,y in texts:
        b = nearest(x,y)
        if not b: continue
        # dims
        d = dim_pat.search(txt)
        if d:
            beam_data[b]["WIDTH"], beam_data[b]["DEPTH"] = d.groups()
        # level
        lv = lvl_pat.search(txt)
        if lv and "+" in txt:
            beam_data[b]["LEVEL"] = lv.group(1)
        # other patterns
        for col,pat in patterns.items():
            if pat.search(txt):
                beam_data[b][col] = pat.search(txt).group(0)

    # Write
    row = 2
    for b in sorted(beam_data):
        data = beam_data[b]
        ws.cell(row=row, column=1, value=data.get("BEAM NO"))
        for idx,h in enumerate(headers[1:], start=2):
            ws.cell(row=row, column=idx, value=data.get(h, "N/A"))
        row += 1

    # Alignment
    for r in ws.iter_rows(min_row=2, max_row=row-1, max_col=len(headers)):
        for c in r: c.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(excel_path)
    print(f"Extracted {row-2} beams to {excel_path}")

if __name__ == '__main__':
    extract_rcc_data('TRIAL1_BS.dxf', 'RCC_Beam_Data_Clean.xlsx', debug=True)
