import ezdxf
import openpyxl
import re
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict

def extract_rcc_data(dxf_path, excel_path, proximity_thresh=5000):
    # Load DXF document
    try:
        doc = ezdxf.readfile(dxf_path)
    except (IOError, ezdxf.DXFStructureError) as e:
        print(f"Error reading DXF: {e}")
        return
    msp = doc.modelspace()

    # Prepare Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beam Data"
    headers = ["BEAM NO", "WIDTH", "DEPTH", "LEVEL", "LEFT BOTTOM", "BOTTOM LEFT AT (DIST)", 
        "MID BOTTOM", "CURTAIL AT ( DIST )", "RIGHT BOTTOM", "BOTTOM RIGHT AT ( DIST )", 
        "BENT UP", "LEFT TOP", "LEFT AT ( DIST )", "MID TOP", "RIGHT TOP", "RIGHT AT ( DIST)", 
        "SFR", "SHEAR STIRUPPS LEG", "SHEAR STIRRUPS DIA ( L )", "LEFT SPACE STIRRUPS", 
        "SHEAR STIRRUPS DIA ( M )", "MID SPACE STIRRUPS", "SHEAR STIRRUPS DIA ( R )", 
        "RIGHT SPACE STIRRUPS", "SHEAR STIRRUP NUMBER", "EXTRA STIRRUP NUMBER", 
        "EXTRA STIRRUP DIA", "HORI LINK DIA", "STIRRUPSID CONTINUOUS END", 
        "DISCONTINUOUS END", "ATTACH MASTER ID"]
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    font = Font(bold=True)
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx, value=h)
        c.fill, c.font = fill, font
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 18

    # Patterns
    beam_no_pattern = re.compile(r"\bB\d+[A-Za-z]*\b")
    section_dim_pattern = re.compile(r"(\d+)\s*[xXx]\s*(\d+)")
    level_pattern = re.compile(r"(\d+\.?\d*)\s*[mM]\b")
    rebar_pattern = re.compile(r"[ØøΦ]\s*(\d+)")
    stirrup_pattern = re.compile(r"(\d+T\d+@\s*\d+)")

    # Collect text
    beam_data = defaultdict(lambda: {"rebars": [], "stirrups": []})
    texts = []
    for ent in msp:
        if ent.dxftype() in ("TEXT", "MTEXT"):
            txt = ent.dxf.text.strip()
            pos = (ent.dxf.insert.x, ent.dxf.insert.y)
            layer = ent.dxf.layer
            texts.append((txt, pos, layer))

    # Process
    for txt, pos, layer in texts:
        # Beam identifier + inline section dims
        bm = beam_no_pattern.search(txt)
        if bm:
            beam = bm.group(0)
            data = beam_data[beam]
            data['position'] = pos
            sec = section_dim_pattern.search(txt)
            if sec:
                data['width'], data['depth'] = sec.groups()
            continue
        # Section dims separate
        sd = section_dim_pattern.search(txt)
        if sd:
            w, d = sd.groups()
            # find nearest beam
            nearest, md = None, float('inf')
            for b, data in beam_data.items():
                if 'position' not in data: continue
                dx, dy = pos[0]-data['position'][0], pos[1]-data['position'][1]
                dist = (dx*dx+dy*dy)**0.5
                if dist < md:
                    nearest, md = b, dist
            if nearest and md < proximity_thresh:
                beam_data[nearest]['width'] = w
                beam_data[nearest]['depth'] = d
            continue
        # Elevation
        lv = level_pattern.search(txt)
        if lv:
            lvl = lv.group(1)
            nearest, md = None, float('inf')
            for b, data in beam_data.items():
                if 'position' not in data: continue
                dx, dy = pos[0]-data['position'][0], pos[1]-data['position'][1]
                dist = (dx*dx+dy*dy)**0.5
                if dist < md:
                    nearest, md = b, dist
            if nearest and md < proximity_thresh:
                beam_data[nearest]['level'] = lvl
            continue
        # Rebars
        rb = rebar_pattern.search(txt)
        if rb:
            size = rb.group(1)
            nearest, md = None, float('inf')
            for b, data in beam_data.items():
                if 'position' not in data: continue
                dx, dy = pos[0]-data['position'][0], pos[1]-data['position'][1]
                dist = (dx*dx+dy*dy)**0.5
                if dist < md:
                    nearest, md = b, dist
            if nearest and md < proximity_thresh:
                beam_data[nearest]['rebars'].append(size)
            continue
        # Stirrups
        st = stirrup_pattern.search(txt)
        if st:
            info = st.group(1)
            nearest, md = None, float('inf')
            for b, data in beam_data.items():
                if 'position' not in data: continue
                dx, dy = pos[0]-data['position'][0], pos[1]-data['position'][1]
                dist = (dx*dx+dy*dy)**0.5
                if dist < md:
                    nearest, md = b, dist
            if nearest and md < proximity_thresh:
                beam_data[nearest]['stirrups'].append(info)
            continue

    # Write out
    r = 2
    for b, data in sorted(beam_data.items()):
        ws.cell(row=r, column=1, value=b)
        ws.cell(row=r, column=2, value=data.get('width', 'N/A'))
        ws.cell(row=r, column=3, value=data.get('depth', 'N/A'))
        ws.cell(row=r, column=4, value=data.get('level', 'N/A'))
        ws.cell(row=r, column=5, value=', '.join(data['rebars']) or 'N/A')
        ws.cell(row=r, column=6, value=', '.join(data['stirrups']) or 'N/A')
        r += 1
    # Center align
    for row in ws.iter_rows(min_row=2, max_row=r-1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Save
    try:
        wb.save(excel_path)
        print(f"Saved {len(beam_data)} beams -> {excel_path}")
    except PermissionError:
        print("Permission denied. Close the file and retry.")

if __name__ == '__main__':
    extract_rcc_data('TRIAL1_BS.dxf', 'RCC_Beam_Data_Enhanced.xlsx')
